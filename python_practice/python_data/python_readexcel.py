from openpyxl import Workbook
from openpyxl import load_workbook


class PracticeExcel:
    def create(self,filename):
        # from openpyxl import Workbook
        # wk = Workbook()

        wk = Workbook()
        ws = wk.active
        ws['A1'] = "身高"
        ws['B1'] = "体重"

        # ws.append(["170cm", "65kg"])
        # ws.append(["182cm", "75kg"])

        self.height = [180, 160, 170, 155]
        self.weight = [66, 50, 40, 30]

        for i in range(len(self.height)):
            ws.append([self.height[i], self.weight[i]])

            # ws.cell(row=2 + i, column=1, value=height[i])
            # ws.cell(row=2 + i, column=2, value=weight[i])

        wk.save(filename)

    def health_check(self,filename,newfilename):
        ld = load_workbook(filename=filename)
        sheet = ld.active
        # sheet = ld.get_sheet_by_name('Sheet')
        sheet["C1"] = "备注"

        for i in range(len(self.height)):
            height = sheet["A" + str(2 + i)].value
            print(height)
            weight = sheet["B" + str(2 + i)].value
            print(weight)
            tmp = (int(height)-70)*0.6
            if weight == tmp:
                sheet.cell(row = 2+i, column = 3, value = "健康")
            else:
                sheet.cell(row = 2+i, column = 3, value = "不健康")
        ld.save(newfilename)

wk = PracticeExcel()
wk.create("new.xlsx")
wk.health_check("new.xlsx","new-updated.xlsx")
