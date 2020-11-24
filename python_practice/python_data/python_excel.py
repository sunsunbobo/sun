# from openpyxl import Workbook
# wk = Workbook()

import openpyxl
wk = openpyxl.Workbook()
ws = wk.active
ws['A1'] = "身高"
ws['B1'] = "体重"

# ws.append(["170cm", "65kg"])
# ws.append(["182cm", "75kg"])

height = [180, 160, 170, 155]
weight = [60, 50, 40, 30]

for i in range(len(height)):
    ws.append([height[i],weight[i]])

    # ws.cell(row=2 + i, column=1, value=height[i])
    # ws.cell(row=2 + i, column=2, value=weight[i])

wk.save("info.xlsx")